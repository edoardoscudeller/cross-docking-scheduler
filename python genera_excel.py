#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import math
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "results_v0.1.txt"
OUTPUT_FILE = "Cross-Docking-Results-v0.1.xlsx"

SCENARIOS = ["uniform", "sparse", "clustered", "asymmetric", "congested", "urgent", "mixed"]

COLUMN_ORDER = [
    "Istanza",
    "Seed",
    "n (inbound)",
    "m (outbound)",
    "Inbound doors",
    "Outbound doors",
    "Lower Bound",
    "Makespan",
    "Gap (%)",
    "CPU Time (s)",
    "CPU Time (min)",
    "release time",
    "Inbound sequence",
    "Outbound sequence",
]

SECTION_TO_COLUMN = {
    "Instance name": "Istanza",
    "Seed": "Seed",
    "N inbound": "n (inbound)",
    "M outbound": "m (outbound)",
    "Inbound doors": "Inbound doors",
    "Outbound doors": "Outbound doors",
    "Lower Bound": "Lower Bound",
    "Makespan": "Makespan",
    "Gap (%)": "Gap (%)",
    "CPU Time (s)": "CPU Time (s)",
    "CPU Time (min)": "CPU Time (min)",
    "Release time": "release time",
    "Inbound sequence": "Inbound sequence",
    "Outbound sequence": "Outbound sequence",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
ALT_FILL = PatternFill("solid", fgColor="F7FBFF")
THIN_GRAY = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

NUMERIC_COLUMNS = {
    "Seed",
    "n (inbound)",
    "m (outbound)",
    "Inbound doors",
    "Outbound doors",
    "Lower Bound",
    "Makespan",
    "Gap (%)",
    "CPU Time (s)",
    "CPU Time (min)",
}

COLUMN_WIDTHS = {
    "Istanza": 30,
    "Seed": 10,
    "n (inbound)": 12,
    "m (outbound)": 13,
    "Inbound doors": 14,
    "Outbound doors": 15,
    "Lower Bound": 14,
    "Makespan": 12,
    "Gap (%)": 12,
    "CPU Time (s)": 14,
    "CPU Time (min)": 16,
    "release time": 42,
    "Inbound sequence": 55,
    "Outbound sequence": 55,
}


def clean_value(value, column_name):
    value = value.strip()

    if value == "":
        return None

    if column_name in {
        "Seed",
        "n (inbound)",
        "m (outbound)",
        "Inbound doors",
        "Outbound doors",
        "Lower Bound",
        "Makespan",
    }:
        try:
            return int(value)
        except ValueError:
            return value

    if column_name in {"Gap (%)", "CPU Time (s)", "CPU Time (min)"}:
        tmp = value.replace("%", "").strip()
        try:
            return float(tmp)
        except ValueError:
            return value

    return value


def parse_blocks(lines):
    scenarios_data = {}
    i = 0
    n = len(lines)

    while i < n:
        line = lines[i].strip()

        if line == "========================================":
            if i + 1 < n and lines[i + 1].strip().startswith("SCENARIO:"):
                scenario = lines[i + 1].strip().split("SCENARIO:", 1)[1].strip()
                i += 3

                sections = {}
                while i < n:
                    cur = lines[i].rstrip("\n").strip()

                    if cur == "========================================":
                        break

                    if cur in SECTION_TO_COLUMN:
                        section_name = cur
                        i += 1
                        values = []

                        while i < n:
                            nxt = lines[i].rstrip("\n")
                            nxt_stripped = nxt.strip()

                            if nxt_stripped == "":
                                i += 1
                                break

                            if nxt_stripped in SECTION_TO_COLUMN:
                                break

                            if nxt_stripped == "========================================":
                                break

                            values.append(nxt.strip())
                            i += 1

                        sections[section_name] = values
                    else:
                        i += 1

                scenarios_data[scenario] = sections
                continue

        i += 1

    return scenarios_data


def build_rows(sections):
    max_len = 0
    for sec_name in SECTION_TO_COLUMN:
        vals = sections.get(sec_name, [])
        max_len = max(max_len, len(vals))

    rows = []
    for idx in range(max_len):
        row = {}
        for sec_name, col_name in SECTION_TO_COLUMN.items():
            vals = sections.get(sec_name, [])
            val = vals[idx] if idx < len(vals) else ""
            row[col_name] = clean_value(val, col_name)
        rows.append(row)

    return rows


def style_sheet(ws, rows_count):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_ORDER))}{max(1, rows_count + 1)}"
    ws.sheet_view.showGridLines = True

    for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 18)

    for row_idx in range(2, rows_count + 2):
        is_alt = (row_idx % 2 == 0)

        for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER

            if is_alt:
                cell.fill = ALT_FILL

            if col_name in NUMERIC_COLUMNS:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

            if col_name in {"CPU Time (s)", "CPU Time (min)"} and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"

            elif col_name == "Gap (%)" and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

            elif col_name in {"Lower Bound", "Makespan", "Seed", "n (inbound)", "m (outbound)", "Inbound doors", "Outbound doors"} and isinstance(cell.value, (int, float)):
                cell.number_format = "0"

    ws.row_dimensions[1].height = 24

    for row_idx in range(2, rows_count + 2):
        max_len = 1
        for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            length = len(str(val))
            if col_name in {"release time", "Inbound sequence", "Outbound sequence"}:
                max_len = max(max_len, min(4, math.ceil(length / 40)))
        ws.row_dimensions[row_idx].height = 18 * max_len


def main():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise FileNotFoundError(f"File non trovato: {INPUT_FILE}")

    text = input_path.read_text(encoding="utf-8")
    lines = text.splitlines()

    scenarios_data = parse_blocks(lines)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for scenario in SCENARIOS:
        ws = wb.create_sheet(title=scenario)

        for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        sections = scenarios_data.get(scenario, {})
        rows = build_rows(sections)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        style_sheet(ws, len(rows))

    wb.save(OUTPUT_FILE)
    print(f"Creato: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()