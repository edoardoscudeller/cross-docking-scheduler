#!/usr/bin/env python3
"""
export_results.py — Cross-Docking Scheduler v0.2
Genera un file Excel formattato con i risultati delle istanze.

Uso:
    python3 export_results.py [risultati_v01.txt] [risultati_v02.txt]

Oppure legge da stdin se i file non sono specificati.
Genera: CD_Results.xlsx  con un foglio per versione (v0.1, v0.2).

Formato atteso dei file di input (uno per riga):
    <istanza>  <makespan>  <cpu_time_s>
    es:  cd_n004_m003   57   0.000123

Stile basato su Transport Analysis sheet (G3-Edoardo-Rohan-Federico.xlsx):
    Header : fill=#70A8E0, bold, font Aptos Narrow 11, altezza riga 34
    Dati   : fill=#B8D3EF, font Aptos Narrow 11, altezza riga 15
"""

import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ─── Stile (Transport Analysis palette) ────────────────────────────────────
FONT_NAME   = "Aptos Narrow"
FONT_SIZE   = 11
HDR_FILL    = PatternFill("solid", fgColor="70A8E0")
ROW_FILL    = PatternFill("solid", fgColor="B8D3EF")
ALT_FILL    = PatternFill("solid", fgColor="D0E6F5")   # azzurro più chiaro x alternanza

HDR_FONT    = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
DATA_FONT   = Font(name=FONT_NAME, size=FONT_SIZE)

CENTER  = Alignment(horizontal="center", vertical="center")
LEFT    = Alignment(horizontal="left",   vertical="center", indent=1)
RIGHT   = Alignment(horizontal="right",  vertical="center")

HDR_ROW_H   = 34
DATA_ROW_H  = 15
TITLE_ROW_H = 22

thin = Side(border_style="thin", color="A0B8CC")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

# ─── Helper ─────────────────────────────────────────────────────────────────
def style_cell(cell, fill, font, align, border=None, num_fmt=None):
    cell.fill      = fill
    cell.font      = font
    cell.alignment = align
    if border:  cell.border = border
    if num_fmt: cell.number_format = num_fmt

def write_header(ws, row, cols):
    for c_idx, label in enumerate(cols, start=2):
        cell = ws.cell(row=row, column=c_idx, value=label)
        style_cell(cell, HDR_FILL, HDR_FONT, CENTER, BORDER)
    ws.row_dimensions[row].height = HDR_ROW_H

def write_data_row(ws, row, values, alt=False):
    fill = ALT_FILL if alt else ROW_FILL
    for c_idx, (val, align, num_fmt) in enumerate(values, start=2):
        cell = ws.cell(row=row, column=c_idx, value=val)
        style_cell(cell, fill, DATA_FONT, align, BORDER, num_fmt)
    ws.row_dimensions[row].height = DATA_ROW_H

# ─── Colonne del report ──────────────────────────────────────────────────────
COLUMNS = [
    "Istanza",
    "n (inbound)",
    "m (outbound)",
    "Makespan",
    "CPU Time (s)",
]
# Larghezze colonne (come nel foglio di riferimento)
COL_WIDTHS = [28, 14, 16, 14, 16]

def parse_instance_name(name):
    """Estrae n e m dal nome istanza es: cd_n004_m003 -> (4, 3)"""
    try:
        parts = name.replace(".dzn","").split("_")
        n = int(next(p[1:] for p in parts if p.startswith("n")))
        m = int(next(p[1:] for p in parts if p.startswith("m")))
        return n, m
    except Exception:
        return "-", "-"

def build_sheet(wb, sheet_name, version_label, data_rows):
    """
    data_rows: lista di dict con chiavi: instance, makespan, cpu_time
    """
    ws = wb.create_sheet(title=sheet_name)

    # ── Margine colonna A ──
    ws.column_dimensions["A"].width = 3

    # ── Titolo ──
    ws.merge_cells(f"B2:{get_column_letter(1 + len(COLUMNS))}2")
    title_cell = ws["B2"]
    title_cell.value = f"Cross-Docking Scheduler — Risultati {version_label}"
    title_cell.font  = Font(name=FONT_NAME, size=13, bold=True)
    title_cell.alignment = CENTER
    title_cell.fill  = HDR_FILL
    ws.row_dimensions[2].height = TITLE_ROW_H

    # ── Sottotitolo / data generazione ──
    ws.merge_cells(f"B3:{get_column_letter(1 + len(COLUMNS))}3")
    sub = ws["B3"]
    sub.value = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font  = Font(name=FONT_NAME, size=9, italic=True)
    sub.alignment = CENTER
    ws.row_dimensions[3].height = 14

    # ── Intestazioni tabella ──
    HDR_ROW = 5
    write_header(ws, HDR_ROW, COLUMNS)

    # ── Imposta larghezze colonne ──
    for c_idx, w in enumerate(COL_WIDTHS, start=2):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    # ── Dati ──
    if not data_rows:
        # Riga placeholder se non ci sono dati
        row_idx = HDR_ROW + 1
        note_cell = ws.cell(row=row_idx, column=2,
                            value="— Nessun risultato disponibile —")
        style_cell(note_cell, ROW_FILL, DATA_FONT, CENTER, BORDER)
        ws.merge_cells(f"B{row_idx}:{get_column_letter(1+len(COLUMNS))}{row_idx}")
        ws.row_dimensions[row_idx].height = DATA_ROW_H
        LAST_ROW = row_idx
    else:
        LAST_ROW = HDR_ROW
        for idx, entry in enumerate(data_rows):
            alt     = (idx % 2 == 1)
            row_idx = HDR_ROW + 1 + idx
            name    = entry.get("instance", "")
            n, m    = parse_instance_name(name)
            mkspan  = entry.get("makespan",  "")
            cpu     = entry.get("cpu_time",  "")
            write_data_row(ws, row_idx, [
                (name,  LEFT,   None),
                (n,     CENTER, "#,##0"),
                (m,     CENTER, "#,##0"),
                (mksp if (mksp := mkspan if isinstance(mkspan,str) else int(mkspan)) or True else mksp,
                          RIGHT,  "#,##0"),
                (cpu,   RIGHT,  "0.000000"),
            ], alt=alt)
            LAST_ROW = row_idx

        # ── Riga TOTALI / statistiche ──
        STATS_ROW = LAST_ROW + 2
        ws.row_dimensions[STATS_ROW].height = DATA_ROW_H
        for c_idx in range(2, 2 + len(COLUMNS)):
            cell = ws.cell(row=STATS_ROW, column=c_idx)
            cell.fill   = HDR_FILL
            cell.font   = HDR_FONT
            cell.border = BORDER
            cell.alignment = CENTER
        ws.cell(row=STATS_ROW, column=2).value = "Statistiche"
        # Makespan: col 5 (indice 5)
        mksp_col  = get_column_letter(5)
        cpu_col   = get_column_letter(6)
        data_start = HDR_ROW + 1
        ws.cell(row=STATS_ROW, column=5).value = \
            f"=MIN({mksp_col}{data_start}:{mksp_col}{LAST_ROW})"
        ws.cell(row=STATS_ROW, column=5).number_format = "#,##0"
        ws.cell(row=STATS_ROW, column=5).font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        ws.cell(row=STATS_ROW, column=5).alignment = RIGHT
        ws.cell(row=STATS_ROW, column=5).border = BORDER
        ws.cell(row=STATS_ROW, column=5).fill = HDR_FILL

        ws.cell(row=STATS_ROW, column=6).value = \
            f"=AVERAGE({cpu_col}{data_start}:{cpu_col}{LAST_ROW})"
        ws.cell(row=STATS_ROW, column=6).number_format = "0.000000"
        ws.cell(row=STATS_ROW, column=6).font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        ws.cell(row=STATS_ROW, column=6).alignment = RIGHT
        ws.cell(row=STATS_ROW, column=6).border = BORDER
        ws.cell(row=STATS_ROW, column=6).fill = HDR_FILL

        # label stats
        ws.cell(row=STATS_ROW-1, column=5).value = "Makespan min"
        ws.cell(row=STATS_ROW-1, column=5).font = Font(name=FONT_NAME, size=8, italic=True)
        ws.cell(row=STATS_ROW-1, column=5).alignment = CENTER
        ws.cell(row=STATS_ROW-1, column=6).value = "CPU avg"
        ws.cell(row=STATS_ROW-1, column=6).font = Font(name=FONT_NAME, size=8, italic=True)
        ws.cell(row=STATS_ROW-1, column=6).alignment = CENTER

    # ── Freeze ──
    ws.freeze_panes = f"B{HDR_ROW + 1}"

    return ws


def load_results(filepath):
    """Legge un file di risultati e restituisce lista di dict."""
    rows = []
    if not os.path.exists(filepath):
        return rows
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) >= 3:
                rows.append({
                    "instance": parts[0],
                    "makespan": int(parts[1]),
                    "cpu_time": float(parts[2]),
                })
            elif len(parts) == 2:
                rows.append({
                    "instance": parts[0],
                    "makespan": int(parts[1]),
                    "cpu_time": 0.0,
                })
    return rows


def main():
    file_v01 = sys.argv[1] if len(sys.argv) > 1 else "results_v01.txt"
    file_v02 = sys.argv[2] if len(sys.argv) > 2 else "results_v02.txt"
    out_file = "CD_Results.xlsx"

    data_v01 = load_results(file_v01)
    data_v02 = load_results(file_v02)

    wb = Workbook()
    # Rimuovi foglio di default
    wb.remove(wb.active)

    build_sheet(wb, "v0.1 — Greedy ERT+SPT",  "v0.1",  data_v01)
    build_sheet(wb, "v0.2 — Greedy Enhanced",  "v0.2",  data_v02)

    wb.save(out_file)
    print(f"[OK] Salvato: {out_file}")


if __name__ == "__main__":
    main()
